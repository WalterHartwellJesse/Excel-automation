import openpyxl
workbook = openpyxl.load_workbook('Máy tồn kho 20252.xlsx') 
sheet = workbook.active
from openpyxl.utils import get_column_letter
for col_num in range(1, 10):  # 1 to 9 for columns A to I
    col_letter = get_column_letter(col_num)
    #print(col_letter, end=' ')
empty_rows = []
for row in range(2,201): 
    is_empty = True
    for col_num in range(1, 10):  # columns A to I
        col_letter = get_column_letter(col_num)
        cell_value = sheet[col_letter + str(row)].value
        if cell_value != None and cell_value != '':
            is_empty = False
            break
    if is_empty:
        empty_rows.append(row)
#print("\nHàng Ngang bị trống:", empty_rows) 
# Find the full big gap block start and end
empty_rows_sorted = sorted(empty_rows)
for i in range(len(empty_rows_sorted) - 1):
    if empty_rows_sorted[i+1] != empty_rows_sorted[i] + 1:
        # Found a gap in empties: small gaps, ignore
        continue
    else:
        # Starting continuous empty rows here
        first_big_empty = empty_rows_sorted[i]
        break
else:
    # No big continuous empty block found
    first_big_empty = empty_rows_sorted[0]
# Find the full big gap block start and end
start_gap = first_big_empty
end_gap = first_big_empty

# Move forward while the next row is still empty
while end_gap + 1 in empty_rows_sorted:
    end_gap += 1

# Now get scattered empty rows outside this big gap
scattered_empty_rows = [r for r in empty_rows_sorted if r < start_gap or r > end_gap]
last_filled_row = first_big_empty - 1
print(f"Hàng mới nhập vào ngày cuối cùng(Hàng trước khi có khoảng trống lớn) là: {last_filled_row}")
print("Bạn muốn chọn hàng ngang nào để sửa?")
print(f"1) Hàng ngang sau lần cuối đã được chỉnh sửa ({last_filled_row + 1})")
print("2) Chọn tùy ý hàng ngang bị thiếu")
choice = input("Chọn 1 hoặc 2: ")

if choice == "1":
    input_row = last_filled_row + 1
else:
    print("Danh sách hàng ngang bị bỏ trống:")
    for r in scattered_empty_rows:
        print(r, end='   ')
    print()  # newline after printing rows
    input_row = int(input("Chọn hàng ngang trong danh sách hàng ngang bị bỏ trống: "))
Date = input("Ngày nhập: ")
Product = input("Sản phẩm: ")
IMEI = input("IMEI: ")
Condition = input("Tình trạng: ")
Purchase_price = input("Giá nhập: ")
Supplier = input("Nhà cung cấp: ")
Accessories = input("Phụ kiện: ")
Selling_price = input("Giá bán: ")
Notes = input("Ghi chú: ")

sheet[f"A{input_row}"] = Date
sheet[f"B{input_row}"] = Product
sheet[f"C{input_row}"] = IMEI
sheet[f"D{input_row}"] = Condition
sheet[f"E{input_row}"] = Purchase_price
sheet[f"F{input_row}"] = Supplier
sheet[f"G{input_row}"] = Accessories
sheet[f"H{input_row}"] = Selling_price
sheet[f"I{input_row}"] = Notes

workbook.save('Máy tồn kho 20252.xlsx')

print(f"Dữ liệu đã được nhập vào hàng {input_row} thành công!")